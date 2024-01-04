import logging
import pathlib
import uuid
import shutil
import os
from openpyxl import load_workbook
from datetime import datetime
from fastapi import APIRouter, Header, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from starlette.responses import FileResponse, StreamingResponse
from typing import List, Optional
from piro_api.database import get_db
import piro_api.processes.catalogs as processes
from piro_api.models.catalogs import CreateProductRequestData, EditProductRequestData
from piro_api import get_settings, AppGenericException
from piro_api.orm import Product, Categoria
from tempfile import NamedTemporaryFile  # Importa NamedTemporaryFile



router = APIRouter()

@router.get('/api/product/table/data')
def get_table_data(limit: int = 10,
                   offset: int = 0,
                   query: str = '',
                   accept_language: str = Header(default='en'),
                   db: Session = Depends(get_db)):
    registers = processes.get_products_catalog_data(db, limit, offset, query)
    data = [processes.make_product_table_from_register(register) for register in registers]

    return data

@router.get('/api/product/autocomplete/data')
def get_table_data(limit: int = 1000,
                   offset: int = 0,
                   query: str = '',
                   accept_language: str = Header(default='en'),
                   db: Session = Depends(get_db)):
    registers = processes.get_products_catalog_data(db, limit, offset, query)
    data = [processes.make_product_table_from_register(register) for register in registers]

    return data

@router.get('/api/product/table/count')
def get_table_count(query: str = '',
                    accept_language: str = Header(default='en'),
                    db: Session = Depends(get_db)):
    return {'count': processes.get_products_catalog_count(db, query)}



@router.post('/api/product')
def create_product(request_data: CreateProductRequestData,
                   accept_language: str = Header(default='en'),
                   db: Session = Depends(get_db)):
    config = get_settings()
    print ("Entro ")
    try:
        product = Product()
        product.nombre = request_data.nombre
        product.precio_menudeo = request_data.precio_menudeo
        product.precio_mayoreo = request_data.precio_mayoreo
        product.stock = request_data.stock
        product.stock_minimo = request_data.stock_minimo
        product.categoria_id = request_data.categoria_id
        product.is_active = request_data.is_active

        db.add(product)
        db.commit()

        return {'id': product.idProducto, 'message': 'Product created'}
    except Exception as e:
        print(e)
        raise AppGenericException(0, 'Error creating product', 400)


@router.put('/api/product/{product_id}')
def edit_product(product_id: int,
                 request_data: EditProductRequestData,
                 accept_language: str = Header(default='en'),
                 db: Session = Depends(get_db)):
    print ("Entro")
    try:
        print (request_data.categoria_id)
        register: Product = db.query(Product).get(product_id)

        register.nombre = request_data.nombre
        register.precio_menudeo = request_data.precio_menudeo
        register.precio_mayoreo = request_data.precio_mayoreo
        register.stock = request_data.stock
        register.stock_minimo = request_data.stock_minimo
        register.categoria_id = request_data.categoria_id
        register.is_active = request_data.is_active

        db.commit()

        return {'message': 'Updated register'}
    except Exception as e:
        raise AppGenericException(0, f'Product can not be updated. {e}', 400)


@router.delete('/api/product/{product_id}')
def delete_product(product_id: int,
                   db: Session = Depends(get_db)):
    try:
        product = db.query(Product).filter(Product.idProducto == product_id).first()
        if product:
            product.is_active = False
            db.commit()
            return {'message': 'Registro eliminado correctamente'}
        else:
            return {'message': 'Producto no encontrado'}
    except Exception as e:
        return {'message': 'Error al eliminar el producto', 'error': str(e)}

@router.get('/api/product/{product_id}')
def get_product_by_id(product_id: int, db: Session = Depends(get_db)):
    product = db.query(Product).filter(Product.idProducto == product_id).first()

    if product is None:
        raise HTTPException(status_code=404, detail='Product not found')

    # Cargar la categoría asociada
    categoria = db.query(Categoria).filter(Categoria.id == product.categoria_id).first()
    product.categoria = categoria

    return product
@router.post('/api/upload-products')
async def upload_products(file: UploadFile = File(...), db: Session = Depends(get_db)):
    try:
        # Guardar el archivo temporal en el disco
        with open('temp.xlsx', 'wb') as temp_file:
            shutil.copyfileobj(file.file, temp_file)

        # Cargar el archivo Excel
        wb = load_workbook(filename='temp.xlsx')
        sheet = wb.active

        # Iterar sobre cada fila del archivo Excel y crear instancias de Product
        for row in sheet.iter_rows(min_row=2, values_only=True):
            nombre, stock, stock_minimo, precio_proveedor, precio, precio_mayoreo, categoria_id = row
            existing_product = db.query(Product).filter_by(nombre=nombre).first()
            if existing_product:
                # Si el producto existe, actualiza la cantidad
                print ("Entro aqui")
                existing_product.stock += stock
            else:
            # Crear una instancia de Product y agregarla a la sesión para la base de datos
                nuevo_producto = Product(
                    nombre=nombre,
                    stock=stock,
                    stock_minimo=stock_minimo,
                    #precio_proveedor=precio_proveedor,
                    precio_menudeo=precio,
                    precio_mayoreo=precio_mayoreo,
                    categoria_id=categoria_id
                )
                db.add(nuevo_producto)

        # Guardar los cambios en la base de datos
        db.commit()

        # Eliminar el archivo temporal
        os.remove('temp.xlsx')

        return {"message": "Datos cargados exitosamente"}

    except Exception as e:
        return {"detail": str(e)}